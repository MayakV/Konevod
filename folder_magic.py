import ctypes
import json
from ctypes import POINTER, Structure, c_wchar, c_int, sizeof, byref
from ctypes.wintypes import BYTE, WORD, DWORD, LPWSTR, LPSTR

from time import sleep
from configparser import RawConfigParser
from openpyxl_image_loader import SheetImageLoader
import os
import shutil
import openpyxl
from pathlib import Path

config = RawConfigParser()

path_to_products_to_add = Path('products_to_add/')


def create_folders(file_path, names):
    clear_folder(file_path)
    for folder_name in names:
        os.mkdir(file_path / folder_name)


# TODO: it can't delete some folders, probably because they hve desktop.ini and windows considers
#       it a system file
def clear_folder(path):
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))


def set_logo(product_code, img):
    folder_path = path_to_products_to_add / product_code
    thumbnail_pth = path_to_products_to_add / product_code / 'thumbnail.jpg'
    img.save(thumbnail_pth)
    config['ViewState'] = {'Mode': '',
                           'Vid': '',
                           'FolderType': 'Pictures',
                           'Logo': thumbnail_pth}
    with open(folder_path / 'desktop.ini', 'w') as desk:
        config.write(desk)


def parse_products(workbook, sheet_name):
    sheet = workbook[sheet_name]

    # range(5, sheet.max_row+1) actual product codes start from fifth row
    # "{}{}".format('A', row) - this gives cell name like A5. Product codes are in A column
    products = {sheet["{}{}".format('A', row)].value: {} for row in range(5, sheet.max_row + 1)}

    if not products:
        raise ValueError("There are no products in the catalog")

    return products


def parse_product_info(workbook: openpyxl.Workbook, sheet_name: str, products: dict) -> dict:
    sheet = workbook[sheet_name]
    for row in range(2, sheet.max_row + 1):
        if sheet["A{}".format(row)].value in products:
            products[sheet["A{}".format(row)].value].update({'Code': sheet["A{}".format(row)].value,
                                                             'Name': sheet["D{}".format(row)].value,
                                                             'Supplier': sheet["W{}".format(row)].value,
                                                             'Material': sheet["X{}".format(row)].value,
                                                             'Price': sheet["AB{}".format(row)].value})
    return products


def parse_photos(workbook, sheet_name, products):
    sheet = workbook[sheet_name]
    image_loader = SheetImageLoader(sheet)

    for row in range(2, sheet.max_row):
        if sheet["A{}".format(row)].value in products:
            image = image_loader.get('B{}'.format(row))
            image = image.convert('RGB')
            products[sheet["A{}".format(row)].value].update({'Image': image})

    return products


def serialize(product, exclude, folder_path='') -> None:
    if not folder_path:
        folder_path = path_to_products_to_add / product["Code"]
    with open(folder_path / 'product_info.json', 'w', encoding='utf-8') as f:
        json.dump({x: product[x] for x in product if x not in exclude}, f)


def deserialize(folder_path) -> dict:
    with open(folder_path / 'product_info.json', 'r', encoding='utf-8') as f:
        return json.load(f)

